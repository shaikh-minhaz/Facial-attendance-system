#attendance.py
from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
from tkinter import messagebox
import os
from tkinter import filedialog
import openpyxl

mydata = []

class Attendance:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1530x790+0+0")
        self.root.title("ATTENDANCE")

        self.var_atten_id = StringVar()
        self.var_atten_roll = StringVar()
        self.var_atten_name = StringVar()
        self.var_atten_dep = StringVar()

        #first img
        img = Image.open(r"C:\Users\Macbook Air\Anti Spoofing\imgs\pp.jpg")
        img = img.resize((1440, 170))
        self.photoimg= ImageTk.PhotoImage(img)

        f_lbl = Label(self.root, image = self.photoimg)
        f_lbl.place(x=0, y=0, width=1440, height=170)


        #bg image
        img3 = Image.open(r"C:\Users\Macbook Air\Anti Spoofing\imgs\tr1.jpg")
        img3 = img3.resize((1530, 710))
        self.photoimg3 = ImageTk.PhotoImage(img3)

        bg_img =Label(self.root, image=self.photoimg3)
        bg_img.place(x=0, y=150, width=1530, height=710)

        heading = Label(bg_img, text="ATTENDANCE RECORDS", font=("Exotc350 DmBd BT", 35, "bold"), bg="black", fg="white")
        heading.place(x=0, y=10, width=1435, height=70)

        main_frame = Frame(bg_img, bd=2, bg="black")
        main_frame.place(x=0, y=90, width=1440, height=600)

        # left label frame
        Left_frame = LabelFrame(main_frame, bd=2, bg="red", relief=RIDGE, text="Student Attendance Details", font=("times new roman", 13, 'bold'))
        Left_frame.place(x=0, y=0, width=730, height=580)

        img_left = Image.open(r"C:\Users\Macbook Air\Anti Spoofing\imgs\art.jpg")
        img_left = img_left.resize((715, 200))
        self.photoimg_left = ImageTk.PhotoImage(img_left)

        f_lbl = Label(Left_frame, image=self.photoimg_left)
        f_lbl.place(x=5, y=0, width=715, height=200)

        left_inside_frame = Frame(Left_frame, bd=2, relief=RIDGE, bg='purple')
        left_inside_frame.place(x=0, y=210, width=720, height=370)


        attendanceId_label = Label(left_inside_frame,text= "ROLL NO:",font = ("times new roman",12,"bold"),bg= 'purple',fg ='yellow')
        attendanceId_label.grid(row= 1, column= 0, padx= 10, pady= 4, sticky= W)

        attendanceID_entry = ttk.Entry(left_inside_frame, width= 20, textvariable= self.var_atten_id,font=('comicsansns',13,'bold'))
        attendanceID_entry.grid(row =1, column= 1, padx= 10, pady= 4, sticky= W)

        #Name
        namlabel = Label(left_inside_frame, text="NAME:", bg="purple", font=("times new roman" ,12," bold"), fg='yellow')
        namlabel.grid(row=1, column=2, padx=4, pady=7)

        name_entry = ttk.Entry(left_inside_frame, width=28, textvariable=self.var_atten_roll, font=("comicsansns", 11," bold"))
        name_entry.grid(row=1, column=3, pady=7)

        #date
        dateLabel = Label(left_inside_frame, text="DATE:", bg="purple", fg='yellow', font=("times new roman", 12," bold"))
        dateLabel.grid(row=2, column=0)

        d_ent = ttk.Entry(left_inside_frame, width=22,textvariable=self.var_atten_name, font=("comicsansns", 11," bold"))
        d_ent.grid(row=2, column=1, pady=7)

        # Time
        timelabel =Label(left_inside_frame, text="TIME:", bg="purple", fg='yellow', font=("times new roman", 12," bold"))
        timelabel.grid(row=2, column=2)
        atten_dep = ttk.Entry(left_inside_frame, width=28, textvariable=self.var_atten_dep, font=("comicsansns", 11," bold"))
        atten_dep.grid(row=2, column=3, pady=7)





        # buttons

        icsv_btn = Button(left_inside_frame, text="Import csv", command=self.importCsv, width=10,font=("times new roman", 13, "bold"), bg="blue", fg="white")
        icsv_btn.grid(row=24, column=1, pady=50)

        ecsv_btn = Button(left_inside_frame, text="Export csv", command=self.exportCsv, width=10, font=("times new roman", 13, "bold"), bg="blue", fg="white")
        ecsv_btn.grid(row=24, column=3)

        bck_btn = Button(left_inside_frame, text="Back", command=self.end_program, width=10, font=("times new roman", 13, "bold"), bg="blue", fg="white")
        bck_btn.grid(row=50, column=2)


        #right label frame
        Right_frame = LabelFrame(main_frame, bd=2, bg="red", relief=RIDGE, text=" Attendance Details", font=("times new roman", 13, 'bold'))
        Right_frame.place(x=800, y=0, width=2000, height=580)

        table_frame = Frame(Right_frame, bd=2, relief=RIDGE, bg="orange")
        table_frame.place(x=10, y=0, width=600, height=520)

        scroll_x = ttk.Scrollbar(table_frame, orient=HORIZONTAL)
        scroll_y = ttk.Scrollbar(table_frame, orient=VERTICAL)

        self.AttendaceReportTable = ttk.Treeview(table_frame, column=( "roll", "name", "date", "time"), xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)

        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)

        scroll_x.config(command=self.AttendaceReportTable.xview)
        scroll_y.config(command=self.AttendaceReportTable.yview)


        self.AttendaceReportTable.heading("roll", text="")
        self.AttendaceReportTable.heading("name", text="")
        self.AttendaceReportTable.heading("date", text="")
        self.AttendaceReportTable.heading("time", text="")


        self.AttendaceReportTable["show"] = "headings"


        self.AttendaceReportTable.column("roll", width=10)
        self.AttendaceReportTable.column("name", width=10)
        self.AttendaceReportTable.column("date", width=10)
        self.AttendaceReportTable.column("time", width=10)



        self.AttendaceReportTable.pack(fill=BOTH, expand=1)

        self.AttendaceReportTable.bind("<ButtonRelease>",self.get_cursor)

    def fetchData(self, rows):
        for i in rows:
            self.AttendaceReportTable.insert("", END, values=i)

    def importCsv(self):
        global mydata

        fln = filedialog.askopenfilename(initialdir=os.getcwd(), title="open xlsx",filetypes=(("Excel Files", ".xlsx"), ("ALL Files",".*")), parent= self.root)
        wb = openpyxl.load_workbook(fln)
        ws = wb.active
        for row in ws.iter_rows(values_only= True):
                mydata.append(row)
        self.fetchData(mydata)

    def end_program(self):
        self.root.destroy()

    def exportCsv(self):
        try:
            if len(mydata) < 1:
                messagebox.showerror("No Data", "No Data Found To Export", parent=self.root)
                return False
            fln = filedialog.asksaveasfilename(initialdir=os.getcwd(), title="Save as xlsx",filetypes=(("Excel File", ".xlsx"), ("ALL Files", ".*")),parent= self.root)
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in mydata:
              ws.append(row)
              wb.save(fln)
            messagebox.showinfo("Data Export", "Your data exported successfully")
        except Exception as es:
            messagebox.showerror("Error", f"Due To: {str(es)}", parent=self.root)

    def get_cursor(self, event=''):
        cursor_row = self.AttendaceReportTable.focus()
        content = self.AttendaceReportTable.item(cursor_row)
        rows = content['values']
        self.var_atten_id.set(rows[0])
        self.var_atten_roll.set(rows[1])
        self.var_atten_name.set(rows[2])
        self.var_atten_dep.set(rows[3])
        self.var_atten_time.set(rows[4])
        self.var_atten_date.set(rows[5])
        self.var_atten_attendance.set(rows[6])

    def reset_data(self):
        self.var_atten_id.set("")
        self.var_atten_roll.set("")
        self.var_atten_name.set("")
        self.var_atten_dep.set("")
        self.var_atten_time.set("")
        self.var_atten_date.set("")
        self.var_atten_attendance.set("")

if __name__ == "__main__":
    root = Tk()
    obj = Attendance(root)
    root.mainloop()
